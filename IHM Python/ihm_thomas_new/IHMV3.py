# ---------------------------------HEADER--------------------------------------
#                          -*- coding: utf-8 -*-
#
#                     Objectives: IHM Robot Infirmier
#
#                 Starting date of programming: 10-2021
#                   ending date of programming: 1-2022
#
#              code by: Paul-Ambroise Burdet, Groupe 6 ESE 2021-2022
#
#      Main program's objective is to be used as a communication interface
#    Between the user and the robot designed during the ESE 2021-2022 project
#
#                                  ToDo:
#          Interface Serial - Zigbee (utiliser et installer PySerial)
#   Stockage des données dans une database (utiliser interface COM de l'ordi)
# -----------------------------------------------------------------------------


# -------------------------------Packages--------------------------------------


from tkinter import * #import everything from Tkinter
from tkinter import ttk #import ttk from tkinter
import os  # provides a portable way of using operating systems
import os.path  # allows to manipulate paths
import sys  # provides access to methods and objects that interacts with conda
import win32com.client  # provides COM connexion
import win32com.client.connect  # provides COM connexion
import serial
from tkinter import messagebox
from openpyxl import load_workbook
#from serial import Serial
from pandas import DataFrame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import time
import math
# -------------------------Classes and Functions-------------------------------





def send_left():
   values = bytearray([4, 101])
   ser.write(values)

def send_right():
   values = bytearray([3, 101])
   ser.write(values)
   
def send_up():
   values = bytearray([1, 101])
   ser.write(values)
   
def send_down():
   values = bytearray([2, 101])
   ser.write(values)
   
def send_stop():
   values = bytearray([0, 101])
   ser.write(values)
   
def send_pwm(pwm):
   print(pwm.get("1.0","end-1c"))
   print(int(pwm.get("1.0","end-1c")))
   print(type(pwm.get("1.0","end-1c")))
   print(type(int(pwm.get("1.0","end-1c"))))
   values = bytearray([5,int(pwm.get("1.0","end-1c")), 101])
   ser.write(values)
   
def send_prescaler(Prescaler):
   print(Prescaler.get("1.0","end-1c"))
   print(int(Prescaler.get("1.0","end-1c")))
   print(type(Prescaler.get("1.0","end-1c")))
   print(type(int(Prescaler.get("1.0","end-1c"))))
   values = bytearray([6,int(Prescaler.get("1.0","end-1c")), 101])
   ser.write(values)
   
def send_database(Nom,Prenom,Adresse,Num):
   filename="hello_world.xlsx"
   wb = load_workbook(filename)
   ws = wb.worksheets[0]
   ws['A1'] = "Nom Patient :"
   ws['B1'] = "Prénom Patient :"
   ws['C1'] = "Adresse Patient :"
   ws['D1'] = "Numéro Patient :"
   ws['E1'] = "Saturation Oxygène :"
   ws['F1'] = "Température :"
   ws['G1'] = "Rythme cardiaque :"
   ws['H1'] = "Intensité du système :"

   i=1
   while ws.cell(i,1).value is not None:
       i=i+1
   ws.cell(i,1).value = Nom.get("1.0","end-1c")
   ws.cell(i,2).value = Prenom.get("1.0","end-1c")
   ws.cell(i,3).value = Adresse.get("1.0","end-1c")
   ws.cell(i,4).value = Num.get("1.0","end-1c")
   wb.save(filename)


# ---------------------------------Main----------------------------------------

lidar_x = []
lidar_y = []

lidar_angles = []
lidar_distances = []

lidar_taille_buffer = 500


ser = serial.Serial('COM19', 115200)


    
window = Tk()
window.title("IHM Projet Robot Infirmier")





def lidar_plt():
    plt.close("all")
    
    global lidar_angles
    global lidar_distances
    
    if len(lidar_angles) > lidar_taille_buffer:
        indice_debut = len(lidar_angles) - lidar_taille_buffer
        lidar_angles    = lidar_angles     [indice_debut:]
        lidar_distances = lidar_distances  [indice_debut:]
    
    
    #fig = plt.figure()
    #ax = fig.add_subplot(111, projection='polar')   
            
    for i in range(len(lidar_angles)) : 
        
        x = lidar_distances[i] * math.cos(2*3.14159*lidar_angles[i]/360)
        y = lidar_distances[i] * math.sin(2*3.14159*lidar_angles[i]/360)
        
        #ax.set_xticks(lidar_angles[i])
        #ax.set_yticks(lidar_distances[i])
        
        plt.plot(x,y, '.')
    
    
    amplitude = 1
    plt.xlim(-amplitude, amplitude)
    plt.ylim(-amplitude, amplitude)
    plt.show()
    
    print("ditances", len(lidar_angles), lidar_distances)
    print("angles", len(lidar_angles), lidar_angles)
    
    
    return


def refresh(Ox_Sanguin_info, Temp_info, Rythme_Card_info, Intensite_info):
    print("Nouvelle boucle refresh")
    print("serial : ", ser.inWaiting())
    j = [2,0,0,0]
    if (ser.inWaiting() >= 4) :
        x = ser.read(ser.inWaiting())
        
        for i in x: # on met en forme pour avoir un tableau lisible
            j.append(i)
    
    #print("j", len(j), j)
    

        
    while len(j) >= 10:

        if (j[0] == 1): # on a une donnee du lidar
            a1 = j[1]
            a2 = j[2]
            d1 = j[3]
            d2 = j[4]

            print(a1, a2, d1, d2)
            
            
            #angle = (a2 * 128 + a1)/64;
            angle = 1.42 * a1;
            
            print("angle",angle)
            distance = (d2 * 255 + d1)/4000
        
            lidar_distances.append(distance)
            lidar_angles.append(angle)
            j = j[5:]
            
        #   print("j", j)
            
            
            
            
            
            
            
        
        if (j[0] == 0):
            Ox_Sanguin_info.config(text = j[1])
            Temp_info.config(text = j[2])
            Rythme_Card_info.config(text = j[3])
            Intensite_info.config(text = j[4])
            j = j[5:]
            
        else :
            j = j[1:]
            
                
                
        
    lidar_plt()
        
        
                        

        
            
   
    window.after(1000, lambda: refresh(Ox_Sanguin_info, Temp_info, Rythme_Card_info, Intensite_info))

    return




menu = Canvas(window,width=2000, height=2000) 
menu.pack_propagate(False)
menu['bg']='#2B50AA'
menu.pack()
menu.create_line(0 , 80 , 1900 , 80 , fill = "white" )
menu.create_line(780 , 82 , 780 , 2000 , fill = "white" )

# Title Creation
Title = Label(text="Informations et Commande du robot", foreground="white")
Title.configure(font=("Ubuntu Light", 36, "roman"))
Title.place(relx=0.5, rely=0.05, anchor=CENTER)
Title['bg']='#2B50AA'

# Window left side
SubTitle = Label(text="Informations du patient", foreground="white")
SubTitle.configure(font=("Ubuntu Light", 24, "roman"))
SubTitle.place(relx=0.25, rely=0.2, anchor=CENTER)
SubTitle['bg']='#2B50AA'

Ox_Sanguin = Label(text="Saturation en oxygène sanguin :", foreground="white")
Ox_Sanguin.configure(font=("Ubuntu Light", 18, "roman"))
Ox_Sanguin.place(relx=0.25, rely=0.30, anchor=CENTER)
Ox_Sanguin['bg']='#2B50AA'

Temp = Label(text="Température du patient :", foreground="white")
Temp.configure(font=("Ubuntu Light", 18, "roman"))
Temp.place(relx=0.25, rely=0.35, anchor=CENTER)
Temp['bg']='#2B50AA'

Rythme_Card = Label(text="Rythme cardiaque du patient :", foreground="white")
Rythme_Card.configure(font=("Ubuntu Light", 18, "roman"))
Rythme_Card.place(relx=0.25, rely=0.40, anchor=CENTER)
Rythme_Card['bg']='#2B50AA'

Nom_Patient_Label = Label(text="Nom du patient :", foreground="white")
Nom_Patient_Label.configure(font=("Ubuntu Light", 18, "roman"))
Nom_Patient_Label.place(relx=0.17, rely=0.50, anchor=CENTER)
Nom_Patient_Label['bg']='#2B50AA'
Nom_Patient = Text(menu, width=20, height=1)
Nom_Patient.configure(font=("Ubuntu Light", 18, "roman"))
Nom_Patient.place(relx=0.32, rely=0.50, anchor=CENTER)

Prenom_Patient_Label = Label(text="Prénom du patient :", foreground="white")
Prenom_Patient_Label.configure(font=("Ubuntu Light", 18, "roman"))
Prenom_Patient_Label.place(relx=0.17, rely=0.55, anchor=CENTER)
Prenom_Patient_Label['bg']='#2B50AA'
Prenom_Patient = Text(menu, width=20, height=1)
Prenom_Patient.configure(font=("Ubuntu Light", 18, "roman"))
Prenom_Patient.place(relx=0.32, rely=0.55, anchor=CENTER)

Adresse_Label = Label(text="Adresse du patient :", foreground="white")
Adresse_Label.configure(font=("Ubuntu Light", 18, "roman"))
Adresse_Label.place(relx=0.17, rely=0.60, anchor=CENTER)
Adresse_Label['bg']='#2B50AA'
Adresse = Text(menu, width=20, height=1)
Adresse.configure(font=("Ubuntu Light", 18, "roman"))
Adresse.place(relx=0.32, rely=0.60, anchor=CENTER)

Num_Label = Label(text="Numéro du patient :", foreground="white")
Num_Label.configure(font=("Ubuntu Light", 18, "roman"))
Num_Label.place(relx=0.17, rely=0.65, anchor=CENTER)
Num_Label['bg']='#2B50AA'
Num = Text(menu, width=20, height=1)
Num.configure(font=("Ubuntu Light", 18, "roman"))
Num.place(relx=0.32, rely=0.65, anchor=CENTER)

Button_Database = Button(menu, text="Envoyer dans base de données", command = lambda: send_database(Nom_Patient,Prenom_Patient,Adresse,Num))
Button_Database.configure(font=("Ubuntu Light", 18, "roman"))
Button_Database.place(relx=0.25, rely=0.75, anchor=CENTER)
Button_Database['bg']='white'

Intensite_Label = Label(text="Intensité du système:", foreground="white")
Intensite_Label.configure(font=("Ubuntu Light", 18, "roman"))
Intensite_Label.place(relx=0.25, rely=0.85, anchor=CENTER)
Intensite_Label['bg']='#2B50AA'

# Window right side
SubTitle = Label(text="Commande du robot", foreground="white")
SubTitle.configure(font=("Ubuntu Light", 24, "roman"))
SubTitle.place(relx=0.75, rely=0.2, anchor=CENTER)
SubTitle['bg']='#2B50AA'

Button_PWM_UP = Button(menu,text = 'AVANCER', command=send_up)
Button_PWM_UP.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM_UP.place(relx=0.75, rely=0.33, anchor=CENTER)
Button_PWM_UP['bg']='white'

Button_PWM_DOWN = Button(menu,text = 'RECULER', command=send_down)
Button_PWM_DOWN.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM_DOWN.place(relx=0.75, rely=0.47, anchor=CENTER)
Button_PWM_DOWN['bg']='white'

Button_PWM_LEFT = Button(menu,text = 'GAUCHE',command=send_left)
Button_PWM_LEFT.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM_LEFT.place(relx=0.67, rely=0.40, anchor=CENTER)
Button_PWM_LEFT['bg']='white'

Button_PWM_RIGHT = Button(menu,text = 'DROITE',command=send_right)
Button_PWM_RIGHT.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM_RIGHT.place(relx=0.825, rely=0.40, anchor=CENTER)
Button_PWM_RIGHT['bg']='white'

Button_PWM_STOP = Button(menu,text = 'STOP',command=send_stop)
Button_PWM_STOP.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM_STOP.place(relx=0.75, rely=0.40, anchor=CENTER)
Button_PWM_STOP['bg']='white'

PWM_Label = Label(text="Nouvelle valeur PWM :", foreground="white")
PWM_Label.configure(font=("Ubuntu Light", 18, "roman"))
PWM_Label.place(relx=0.73, rely=0.65, anchor=CENTER)
PWM_Label['bg']='#2B50AA'
PWM = Text(menu, width=10, height=1)
PWM.configure(font=("Ubuntu Light", 18, "roman"))
PWM.place(relx=0.89, rely=0.65, anchor=CENTER)

Prescaler_Label = Label(text="Nouvelle valeur Prescaler :", foreground="white")
Prescaler_Label.configure(font=("Ubuntu Light", 18, "roman"))
Prescaler_Label.place(relx=0.73, rely=0.60, anchor=CENTER)
Prescaler_Label['bg']='#2B50AA'
Prescaler = Text(menu, width=10, height=1)
Prescaler.configure(font=("Ubuntu Light", 18, "roman"))
Prescaler.place(relx=0.89, rely=0.60, anchor=CENTER)

Button_Prescaler = Button(menu, text="Mise à jour Prescaler",command = lambda: send_prescaler(Prescaler))
Button_Prescaler.configure(font=("Ubuntu Light", 18, "roman"))
Button_Prescaler.place(relx=0.75, rely=0.70, anchor=CENTER)
Button_Prescaler['bg']='white'

Button_PWM = Button(menu, text="Mise à jour PWM",command = lambda: send_pwm(PWM))
Button_PWM.configure(font=("Ubuntu Light", 18, "roman"))
Button_PWM.place(relx=0.75, rely=0.77, anchor=CENTER)
Button_PWM['bg']='white'

data = [0,0,0,0]

Ox_Sanguin_info = Label(text=data[0], foreground="white")
Ox_Sanguin_info.configure(font=("Ubuntu Light", 18, "roman"))
Ox_Sanguin_info.place(relx=0.38, rely=0.30, anchor=CENTER)
Ox_Sanguin_info['bg']='#2B50AA'

Temp_info = Label(text=data[1], foreground="white")
Temp_info.configure(font=("Ubuntu Light", 18, "roman"))
Temp_info.place(relx=0.35, rely=0.35, anchor=CENTER)
Temp_info['bg']='#2B50AA'

Rythme_Card_info = Label(text=data[2], foreground="white")
Rythme_Card_info.configure(font=("Ubuntu Light", 18, "roman"))
Rythme_Card_info.place(relx=0.37, rely=0.40, anchor=CENTER)
Rythme_Card_info['bg']='#2B50AA'

Intensite_info = Label(text=data[3], foreground="white")
Intensite_info.configure(font=("Ubuntu Light", 18, "roman"))
Intensite_info.place(relx=0.34, rely=0.85, anchor=CENTER)
Intensite_info['bg']='#2B50AA'

data = [20,20,30,40]


ser.reset_input_buffer()
ser.reset_output_buffer()


Button_Acq = Button(menu, text="Acquérir",command = lambda: refresh(Ox_Sanguin_info, Temp_info, Rythme_Card_info, Intensite_info))
Button_Acq.configure(font=("Ubuntu Light", 18, "roman"))
Button_Acq.place(relx=0.75, rely=0.85, anchor=CENTER)
Button_Acq['bg']='white'        

window.after(1000, lambda: refresh(Ox_Sanguin_info, Temp_info, Rythme_Card_info, Intensite_info))
window.mainloop()





















ser.close()
print("serial fermé")
plt.close('all')
print("Figures fermees")